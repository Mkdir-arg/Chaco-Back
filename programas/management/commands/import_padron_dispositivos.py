"""Importa un padrón normalizado de Dispositivos y Merenderos."""

import csv
from datetime import date
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from programas.models import Dispositivo, Merendero, TipoDispositivo


class Command(BaseCommand):
    help = "Importa dispositivos y merenderos desde un CSV o XLSX normalizado."

    REQUIRED_COLUMNS = {"entidad", "codigo", "nombre", "domicilio", "responsable_nombre"}

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Archivo CSV o XLSX normalizado.")
        parser.add_argument("--fuente", required=True, help="Fuente del padrón.")
        parser.add_argument("--fecha", required=True, help="Fecha de referencia en formato AAAA-MM-DD.")
        parser.add_argument("--responsable", required=True, help="Responsable de la carga.")

    def handle(self, *args, **options):
        archivo = Path(options["file"])
        if not archivo.is_file():
            raise CommandError(f"No existe el archivo: {archivo}")
        if archivo.suffix.lower() not in {".csv", ".xlsx"}:
            raise CommandError("El importador admite archivos CSV o XLSX normalizados.")
        try:
            fecha = date.fromisoformat(options["fecha"])
        except ValueError as error:
            raise CommandError("La fecha debe tener el formato AAAA-MM-DD.") from error

        fuente = options["fuente"].strip()
        responsable = options["responsable"].strip()
        if not fuente or not responsable:
            raise CommandError("Fuente y responsable son obligatorios.")

        filas = self._leer_filas(archivo)

        creados = {"dispositivos": 0, "merenderos": 0}
        omitidos, errores = 0, []
        with transaction.atomic():
            for numero, fila in enumerate(filas, start=2):
                try:
                    resultado = self._importar_fila(fila, fuente=fuente, fecha=fecha, responsable=responsable)
                except ValueError as error:
                    errores.append(f"Fila {numero}: {error}")
                    continue
                if resultado is None:
                    omitidos += 1
                else:
                    creados[resultado] += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Importación completada: {creados['dispositivos']} dispositivos, "
                f"{creados['merenderos']} merenderos, {omitidos} omitidos por duplicado."
            )
        )
        for error in errores:
            self.stderr.write(self.style.WARNING(error))

    @staticmethod
    def _texto(fila, campo):
        valor = fila.get(campo)
        return "" if valor is None else str(valor).strip()

    def _leer_filas(self, archivo):
        if archivo.suffix.lower() == ".csv":
            with archivo.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                headers = reader.fieldnames or ()
                filas = list(reader)
        else:
            libro = load_workbook(archivo, read_only=True, data_only=True)
            try:
                hoja = libro.active
                valores = hoja.iter_rows(values_only=True)
                headers = [str(valor).strip() if valor is not None else "" for valor in next(valores, ())]
                filas = [dict(zip(headers, fila)) for fila in valores if any(valor is not None for valor in fila)]
            finally:
                libro.close()
        faltantes = self.REQUIRED_COLUMNS - set(headers)
        if faltantes:
            raise CommandError(f"Faltan columnas obligatorias: {', '.join(sorted(faltantes))}.")
        return filas

    def _importar_fila(self, fila, *, fuente, fecha, responsable):
        entidad = self._texto(fila, "entidad").upper()
        codigo = self._texto(fila, "codigo").upper()
        nombre = self._texto(fila, "nombre")
        domicilio = self._texto(fila, "domicilio")
        responsable_nombre = self._texto(fila, "responsable_nombre")
        if not codigo:
            raise ValueError("el código institucional es obligatorio")
        if not nombre or not domicilio or not responsable_nombre:
            raise ValueError("nombre, domicilio y responsable_nombre son obligatorios")

        metadata = {"fuente_padron": fuente, "fecha_padron": fecha, "responsable_padron": responsable}
        if entidad == "DISPOSITIVO":
            tipo_codigo = self._texto(fila, "tipo")
            tipo = TipoDispositivo.objects.filter(codigo__iexact=tipo_codigo).first()
            if tipo is None:
                raise ValueError("el tipo de dispositivo es obligatorio y debe existir")
            return self._crear_si_no_existe(
                Dispositivo,
                codigo,
                {
                    "nombre": nombre,
                    "tipo": tipo,
                    "domicilio": domicilio,
                    "localidad": self._texto(fila, "localidad"),
                    "responsable_nombre": responsable_nombre,
                    "responsable_documento": self._texto(fila, "responsable_documento"),
                    "contacto_telefono": self._texto(fila, "contacto_telefono"),
                    "contacto_email": self._texto(fila, "contacto_email"),
                    "horarios": self._texto(fila, "dias_horarios"),
                    **metadata,
                },
                "dispositivos",
            )
        if entidad == "MERENDERO":
            return self._crear_si_no_existe(
                Merendero,
                codigo,
                {
                    "nombre": nombre,
                    "domicilio": domicilio,
                    "zona": self._texto(fila, "zona"),
                    "barrio": self._texto(fila, "barrio"),
                    "dias_horarios": self._texto(fila, "dias_horarios"),
                    "telefono": self._texto(fila, "contacto_telefono"),
                    "responsable_nombre": responsable_nombre,
                    "responsable_documento": self._texto(fila, "responsable_documento"),
                    "responsable_email": self._texto(fila, "contacto_email"),
                    **metadata,
                },
                "merenderos",
            )
        raise ValueError("entidad debe ser DISPOSITIVO o MERENDERO")

    @staticmethod
    def _crear_si_no_existe(modelo, codigo, defaults, resultado):
        try:
            with transaction.atomic():
                _objeto, creado = modelo.objects.get_or_create(codigo=codigo, defaults=defaults)
        except IntegrityError:
            creado = False
        return resultado if creado else None
