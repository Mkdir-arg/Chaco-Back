"""Exportación común de datasets tabulares a CSV y XLSX."""

import csv
from datetime import datetime

from django.http import HttpResponse, HttpResponseBadRequest
from django.utils import timezone
from openpyxl import Workbook


def celda_segura(valor):
    if isinstance(valor, datetime) and timezone.is_aware(valor):
        return timezone.localtime(valor).replace(tzinfo=None)
    if isinstance(valor, str) and valor.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{valor}"
    return valor


def _nombre_hoja(nombre):
    """Excel limita el nombre de hoja a 31 caracteres y prohíbe ``[]:*?/\\``."""
    limpio = "".join("_" if c in "[]:*?/\\" else c for c in nombre).strip() or "Hoja"
    return limpio[:31]


def respuesta_libro(hojas, nombre, alcance=""):
    """XLSX de varias hojas: ``hojas`` es una lista de ``(nombre_hoja, Reporte)``.

    Cada hoja arranca con el alcance aplicado (RN-16 del análisis #366) para que un
    archivo descargado sea reconstruible, una fila en blanco, y después el reporte.
    Complementa a :func:`respuesta_reporte`, que sigue sirviendo a una hoja sola.
    """
    libro = Workbook(write_only=True)
    for nombre_hoja, reporte in hojas:
        hoja = libro.create_sheet(_nombre_hoja(nombre_hoja))
        if alcance:
            hoja.append([celda_segura(f"Alcance: {alcance}")])
            hoja.append([])
        hoja.append(list(reporte.encabezados))
        for fila in reporte.filas:
            hoja.append([celda_segura(valor) for valor in fila])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre}.xlsx"'
    libro.save(response)
    return response


def respuesta_reporte(reporte, formato, nombre, alcance=""):
    filas = ([celda_segura(valor) for valor in fila] for fila in reporte.filas)
    if formato == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{nombre}.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        if alcance:
            writer.writerow([celda_segura(f"Alcance: {alcance}")])
            writer.writerow([])
        writer.writerow(reporte.encabezados)
        writer.writerows(filas)
        return response
    if formato == "xlsx":
        libro = Workbook(write_only=True)
        hoja = libro.create_sheet("Reporte")
        if alcance:
            hoja.append([celda_segura(f"Alcance: {alcance}")])
            hoja.append([])
        hoja.append(list(reporte.encabezados))
        for fila in filas:
            hoja.append(fila)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{nombre}.xlsx"'
        libro.save(response)
        return response
    return HttpResponseBadRequest("Formato de exportación no válido.")
