"""Padrón de habilitados por Excel (RN-P14, #299; identidad desde el Cambio 57).

Uno por **convocatoria**: lo usan todos sus relevamientos, del link público y
de la app de campo. El operador sube un .xlsx de hasta seis columnas —
``documento, sexo, nombre, apellido, fecha de nacimiento, localidad`` — desde la
convocatoria. Acá viven el parser, la carga (reemplazo total, transaccional),
el chequeo ``esta_habilitado`` que consume el paso 1 del link **antes** de
consultar Base de Personas, la fila por DNI + sexo que alimenta la cascada de
identidad (``programas.services.identidad``) y el cruce automático que valida
los casos pendientes cuando llega un padrón con datos.

Las tres columnas de identidad y la localidad son opcionales por fila: una fila
con solo documento y sexo **habilita pero no valida** (RN-2). Los padrones de
dos columnas cargados antes del cambio siguen valiendo tal cual (RN-7).

Normalización en ambos sentidos: el DNI se reduce a dígitos y el sexo a F/M
(acepta "f", "Femenino", "MASCULINO", etc.), tanto al cargar el Excel como al
chequear lo tipeado en el paso 1. La localidad se cruza por nombre contra el
catálogo (``core.Localidad``) sin acentos ni mayúsculas; si no coincide queda
solo el texto y la carga lo reporta.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction

from programas.models import Convocatoria, Formulario, PadronHabilitado

# Tamaño máximo del Excel (los padrones reales son de cientos de filas).
PADRON_MAX_BYTES = 2 * 1024 * 1024

# Encabezados de la plantilla, en el orden de las columnas.
COLUMNAS = ("documento", "sexo", "nombre", "apellido", "fecha de nacimiento", "localidad")

_SEXOS = {
    "F": "F",
    "FEMENINO": "F",
    "MUJER": "F",
    "M": "M",
    "MASCULINO": "M",
    "HOMBRE": "M",
    "VARON": "M",
    "VARÓN": "M",
}

_ENCABEZADOS_DNI = {"DOCUMENTO", "DNI", "NRO DOCUMENTO", "NUMERO DE DOCUMENTO", "NÚMERO DE DOCUMENTO"}

_FORMATOS_FECHA = ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d")


def normalizar_dni(valor):
    # openpyxl entrega floats (30123456.0) en Excels exportados desde CSV,
    # pandas o LibreOffice; sin este cast el DNI quedaba con un 0 de más.
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return "".join(ch for ch in str(valor or "") if ch.isdigit())


def normalizar_sexo(valor):
    return _SEXOS.get(str(valor or "").strip().upper(), "")


def normalizar_texto(valor):
    """Nombre, apellido o localidad tal como vienen, sin espacios de más."""
    return re.sub(r"\s+", " ", str(valor or "")).strip()


def normalizar_fecha(valor):
    """Devuelve ``(fecha, invalida)``. ``invalida`` es True solo cuando había
    algo escrito que no se pudo interpretar; una celda vacía no es inválida."""
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None, False
    if isinstance(valor, datetime):
        return valor.date(), False
    if isinstance(valor, date):
        return valor, False
    if isinstance(valor, (int, float)):
        # Número de serie de Excel (fechas sin formato de celda).
        try:
            from openpyxl.utils.datetime import from_excel

            convertido = from_excel(valor)
        except (ValueError, TypeError, OverflowError):
            return None, True
        return (convertido.date() if isinstance(convertido, datetime) else convertido), False
    texto = str(valor).strip()
    for formato in _FORMATOS_FECHA:
        try:
            return datetime.strptime(texto, formato).date(), False
        except ValueError:
            continue
    return None, True


def clave_localidad(texto):
    """«Sáenz Peña» y «SAENZ PENA» son la misma localidad."""
    plano = unicodedata.normalize("NFD", str(texto or ""))
    plano = "".join(ch for ch in plano if unicodedata.category(ch) != "Mn")
    plano = re.sub(r"[^a-z0-9]+", " ", plano.lower())
    return plano.strip()


@dataclass
class ResumenPadron:
    """Lo que la carga informa al operador (y lo que devuelve el parser)."""

    validas: int = 0
    con_identidad: int = 0
    rechazadas: int = 0
    fechas_invalidas: int = 0
    localidades_no_reconocidas: list = field(default_factory=list)
    casos_validados: int = 0

    def mensaje(self):
        partes = [f"{self.validas} habilitado{'s' if self.validas != 1 else ''}"]
        partes.append(f"{self.con_identidad} con identidad completa")
        if self.rechazadas:
            partes.append(f"{self.rechazadas} fila{'s' if self.rechazadas != 1 else ''} ignorada{'s' if self.rechazadas != 1 else ''}")
        if self.fechas_invalidas:
            partes.append(f"{self.fechas_invalidas} fecha{'s' if self.fechas_invalidas != 1 else ''} sin interpretar")
        if self.localidades_no_reconocidas:
            partes.append(f"{len(self.localidades_no_reconocidas)} localidad(es) no reconocida(s)")
        if self.casos_validados:
            partes.append(f"{self.casos_validados} caso{'s' if self.casos_validados != 1 else ''} pendiente{'s' if self.casos_validados != 1 else ''} validado{'s' if self.casos_validados != 1 else ''}")
        return "Padrón cargado: " + " · ".join(partes) + "."


def parsear_padron(archivo):
    """Lee el Excel y devuelve ``(entradas, resumen)``.

    ``entradas``: lista de dicts ``{dni, sexo, nombre, apellido,
    fecha_nacimiento, localidad_texto}`` normalizados y sin duplicados (gana la
    primera aparición de cada DNI). ``resumen``: ``ResumenPadron`` con las filas
    rechazadas y las fechas que no se pudieron interpretar (se informan, no
    rompen la carga; la fila queda sin fecha).

    Levanta ``ValidationError`` si el archivo no es un .xlsx legible, pesa de
    más o no aporta ninguna fila válida.
    """
    nombre = getattr(archivo, "name", "") or ""
    if not nombre.lower().endswith(".xlsx"):
        raise ValidationError(
            "El padrón debe ser un archivo Excel (.xlsx) con las columnas: "
            "documento, sexo, nombre, apellido, fecha de nacimiento y localidad."
        )
    if getattr(archivo, "size", 0) > PADRON_MAX_BYTES:
        raise ValidationError("El padrón no puede superar los 2 MB.")

    from openpyxl import load_workbook

    try:
        libro = load_workbook(archivo, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl levanta variantes según el archivo
        raise ValidationError("No se pudo leer el archivo: no es un Excel .xlsx válido.") from exc

    resumen = ResumenPadron()
    try:
        hoja = libro.active
        entradas = []
        vistos = set()
        for indice, fila in enumerate(hoja.iter_rows(min_col=1, max_col=len(COLUMNAS), values_only=True)):
            celdas = (tuple(fila) + (None,) * len(COLUMNAS))[: len(COLUMNAS)]
            crudo_dni, crudo_sexo, crudo_nombre, crudo_apellido, crudo_fecha, crudo_localidad = celdas
            dni = normalizar_dni(crudo_dni)
            sexo = normalizar_sexo(crudo_sexo)
            if not any(str(c or "").strip() for c in celdas):
                continue  # fila totalmente vacía
            if indice == 0 and not dni and str(crudo_dni or "").strip().upper() in _ENCABEZADOS_DNI:
                continue  # fila de encabezado
            if not dni or not sexo or len(dni) not in (7, 8):
                resumen.rechazadas += 1
                continue
            if dni in vistos:
                resumen.rechazadas += 1
                continue
            fecha, invalida = normalizar_fecha(crudo_fecha)
            if invalida:
                resumen.fechas_invalidas += 1
            vistos.add(dni)
            entradas.append(
                {
                    "dni": dni,
                    "sexo": sexo,
                    "nombre": normalizar_texto(crudo_nombre),
                    "apellido": normalizar_texto(crudo_apellido),
                    "fecha_nacimiento": fecha,
                    "localidad_texto": normalizar_texto(crudo_localidad),
                }
            )
    finally:
        libro.close()

    if not entradas:
        raise ValidationError(
            "El padrón no tiene filas válidas. Se espera un .xlsx con documento y sexo (F/M) "
            "y, opcionalmente, nombre, apellido, fecha de nacimiento y localidad."
        )
    resumen.validas = len(entradas)
    resumen.con_identidad = sum(1 for e in entradas if e["nombre"] and e["apellido"])
    return entradas, resumen


def _convocatoria_de(objetivo):
    """El padrón es de la convocatoria; se acepta también un relevamiento y se
    resuelve la suya (compatibilidad con los llamadores del Cambio 41)."""
    if isinstance(objetivo, Convocatoria):
        return objetivo
    convocatoria = getattr(objetivo, "convocatoria", None)
    if convocatoria is None:
        raise TypeError("Se espera una Convocatoria o un Relevamiento.")
    return convocatoria


def _entrada(item):
    """Acepta la tupla ``(dni, sexo)`` histórica o el dict de seis campos."""
    if isinstance(item, dict):
        return {
            "dni": normalizar_dni(item.get("dni")),
            "sexo": normalizar_sexo(item.get("sexo")),
            "nombre": normalizar_texto(item.get("nombre")),
            "apellido": normalizar_texto(item.get("apellido")),
            "fecha_nacimiento": item.get("fecha_nacimiento"),
            "localidad_texto": normalizar_texto(item.get("localidad_texto")),
        }
    dni, sexo = item
    return {
        "dni": normalizar_dni(dni),
        "sexo": normalizar_sexo(sexo),
        "nombre": "",
        "apellido": "",
        "fecha_nacimiento": None,
        "localidad_texto": "",
    }


def _indice_localidades():
    from core.models import Localidad

    indice = {}
    for localidad in Localidad.objects.only("id", "nombre").order_by("id"):
        indice.setdefault(clave_localidad(localidad.nombre), localidad)
    return indice


@transaction.atomic
def cargar_padron(objetivo, archivo, entradas, usuario=None):
    """Reemplaza el padrón de la convocatoria por ``entradas`` (reemplazo total,
    no merge), guarda el Excel original para trazabilidad y **valida los casos
    pendientes** que ahora figuren con nombre y apellido (RN-5). Devuelve el
    ``ResumenPadron`` de la carga."""
    convocatoria = _convocatoria_de(objetivo)
    filas = [_entrada(item) for item in entradas]
    resumen = ResumenPadron(validas=len(filas))
    localidades = _indice_localidades() if any(f["localidad_texto"] for f in filas) else {}

    objetos = []
    for fila in filas:
        localidad = None
        if fila["localidad_texto"]:
            localidad = localidades.get(clave_localidad(fila["localidad_texto"]))
            if localidad is None and fila["localidad_texto"] not in resumen.localidades_no_reconocidas:
                resumen.localidades_no_reconocidas.append(fila["localidad_texto"])
        objetos.append(
            PadronHabilitado(
                convocatoria=convocatoria,
                dni=fila["dni"],
                sexo=fila["sexo"],
                nombre=fila["nombre"],
                apellido=fila["apellido"],
                fecha_nacimiento=fila["fecha_nacimiento"],
                localidad=localidad,
                localidad_texto=fila["localidad_texto"],
            )
        )
    resumen.con_identidad = sum(1 for o in objetos if o.tiene_identidad)

    convocatoria.padron.all().delete()
    PadronHabilitado.objects.bulk_create(objetos)
    if archivo is not None:
        # El parser ya consumió el stream: rebobinar antes de persistirlo.
        if hasattr(archivo, "seek"):
            archivo.seek(0)
        convocatoria.padron_archivo = archivo
        convocatoria.save(update_fields=["padron_archivo", "modificado"])
    resumen.casos_validados = validar_casos_pendientes(convocatoria, usuario)
    return resumen


def esta_habilitado(objetivo, dni, sexo):
    """¿DNI+sexo pueden inscribirse? Sin padrón el link es abierto (RN-P14)."""
    convocatoria = _convocatoria_de(objetivo)
    if not convocatoria.padron.exists():
        return True
    return convocatoria.padron.filter(dni=normalizar_dni(dni), sexo=normalizar_sexo(sexo)).exists()


def fila_padron(objetivo, dni, sexo):
    """La fila del padrón para DNI + sexo, o ``None``."""
    dni, sexo = normalizar_dni(dni), normalizar_sexo(sexo)
    if not dni or not sexo:
        return None
    return _convocatoria_de(objetivo).padron.filter(dni=dni, sexo=sexo).first()


def convocatoria_con_identidad(convocatorias, dni, sexo):
    """Entre varias convocatorias, la primera cuyo padrón tiene a DNI + sexo
    **con** nombre y apellido; ``None`` si ninguna. Una sola consulta (la app
    de campo puede tener varios relevamientos vigentes)."""
    dni, sexo = normalizar_dni(dni), normalizar_sexo(sexo)
    convocatorias = list(convocatorias)
    if not convocatorias or not dni or not sexo:
        return None
    orden = {c.pk: i for i, c in enumerate(convocatorias)}
    filas = (
        PadronHabilitado.objects.filter(convocatoria__in=convocatorias, dni=dni, sexo=sexo)
        .exclude(nombre="")
        .exclude(apellido="")
        .values_list("convocatoria_id", flat=True)
    )
    ids = sorted(set(filas), key=orden.get)
    return next((c for c in convocatorias if c.pk == ids[0]), None) if ids else None


def datos_de_fila(fila):
    """Lo que la fila aporta a la identificación: mismo contrato que la Gran Base
    (nombre, apellido, fecha ISO) más la localidad para el legajo."""
    return {
        "nombre": fila.nombre,
        "apellido": fila.apellido,
        "fecha_nacimiento": fila.fecha_nacimiento.isoformat() if fila.fecha_nacimiento else "",
        "localidad_id": fila.localidad_id,
        "localidad_texto": fila.localidad_texto,
    }


def _identidad_del_caso(formulario):
    if formulario.ciudadano_id:
        return formulario.ciudadano.dni, formulario.ciudadano.genero
    datos = formulario.datos_identificacion if isinstance(formulario.datos_identificacion, dict) else {}
    return datos.get("dni", ""), datos.get("sexo") or datos.get("genero") or ""


def validar_casos_pendientes(convocatoria, usuario=None):
    """Cruce automático (RN-5): los casos de la convocatoria **sin validar** que
    figuran en el padrón con nombre y apellido pasan a validados por padrón.

    Solo toca casos pendientes y no forzados; nunca desvalida. Completa en el
    ciudadano lo que estaba vacío (no pisa lo cargado) y deja traza por caso.
    Devuelve cuántos validó.
    """
    from programas.services.becas import registrar_traza

    if not convocatoria.padron.exists():
        return 0
    pendientes = Formulario.objects.filter(
        relevamiento__convocatoria=convocatoria,
        validado_renaper=False,
        identidad_forzada=False,
    ).select_related("ciudadano")

    validados = 0
    for formulario in pendientes:
        dni, sexo = _identidad_del_caso(formulario)
        fila = fila_padron(convocatoria, dni, sexo) if dni else None
        if fila is None or not fila.tiene_identidad:
            continue
        cambios = [("Validación de identidad", "Pendiente", "Validada por padrón")]
        ciudadano = formulario.ciudadano
        if ciudadano is not None:
            actualizados = []
            for campo, valor in (
                ("nombre", fila.nombre),
                ("apellido", fila.apellido),
                ("fecha_nacimiento", fila.fecha_nacimiento),
                ("localidad", fila.localidad),
            ):
                if valor and not getattr(ciudadano, campo):
                    setattr(ciudadano, campo, valor)
                    actualizados.append(campo)
                    cambios.append((f"Ciudadano · {campo}", "", str(valor)))
            if actualizados:
                ciudadano.save(update_fields=[*actualizados, "modificado"])
        elif isinstance(formulario.datos_identificacion, dict):
            datos = dict(formulario.datos_identificacion)
            for campo, valor in (
                ("nombre", fila.nombre),
                ("apellido", fila.apellido),
                ("fecha_nacimiento", fila.fecha_nacimiento.isoformat() if fila.fecha_nacimiento else ""),
            ):
                if valor and not datos.get(campo):
                    datos[campo] = valor
            if fila.localidad_id and not datos.get("localidad_id"):
                datos["localidad_id"] = fila.localidad_id
            datos["origen"] = "padron"
            formulario.datos_identificacion = datos
        formulario.validado_renaper = True
        formulario.origen_validacion = Formulario.OrigenValidacion.PADRON
        formulario.save(update_fields=["validado_renaper", "origen_validacion", "datos_identificacion", "modificado"])
        registrar_traza(formulario, usuario, cambios)
        validados += 1
    return validados


def plantilla_padron():
    """El .xlsx de ejemplo que se descarga desde la convocatoria: encabezados y
    dos filas —una completa, una con solo documento y sexo—."""
    from openpyxl import Workbook

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Padrón"
    hoja.append(list(COLUMNAS))
    hoja.append(["30123456", "F", "María Luján", "Gómez", "14/03/1991", "Resistencia"])
    hoja.append(["28111222", "M", "", "", "", ""])
    for columna, ancho in zip("ABCDEF", (14, 8, 22, 22, 20, 22)):
        hoja.column_dimensions[columna].width = ancho
    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()
