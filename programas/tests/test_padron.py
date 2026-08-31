"""Tests del padrón de habilitados por Excel (#299, análisis #289; Cambio 57).

Desde el Cambio 57 el padrón es de la **convocatoria** (lo usan el link y la
app), tiene seis columnas y valida la identidad cuando trae nombre y apellido.
La cascada de identidad y el cruce automático se prueban en
``test_padron_identidad``; acá queda el parser, la carga y la pantalla.
"""

from datetime import date
from io import BytesIO, StringIO

from django.contrib.auth.models import Group, User
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse

from core.models import Localidad, Municipio, Provincia
from programas.forms import RelevamientoForm
from programas.management.commands.seed_becas import ROL_ADMIN
from programas.models import Convocatoria, Relevamiento, Segmento
from programas.services.padron import (
    cargar_padron,
    clave_localidad,
    esta_habilitado,
    fila_padron,
    normalizar_fecha,
    parsear_padron,
    plantilla_padron,
)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx(filas, nombre="padron.xlsx"):
    from openpyxl import Workbook

    libro = Workbook()
    hoja = libro.active
    for fila in filas:
        hoja.append(fila)
    buffer = BytesIO()
    libro.save(buffer)
    return SimpleUploadedFile(nombre, buffer.getvalue(), content_type=XLSX_MIME)


def _pares(entradas):
    return [(e["dni"], e["sexo"]) for e in entradas]


class _BasePadronTest(TestCase):
    def setUp(self):
        self.segmento = Segmento.objects.create(nombre="Seg", cupo_maximo=100)
        self.convocatoria = Convocatoria.objects.create(
            nombre="Conv",
            segmento=self.segmento,
            fecha_inicio=date(2026, 1, 1),
            fecha_fin=date(2026, 12, 31),
        )
        self.relevamiento = Relevamiento.objects.create(
            convocatoria=self.convocatoria,
            tipo=Relevamiento.Tipo.PUBLICO,
            fecha_asignada=date(2026, 6, 1),
            fecha_hasta=date(2026, 6, 30),
        )


class ParserPadronTests(TestCase):
    def test_parsea_filas_validas_y_reporta_rechazadas(self):
        archivo = _xlsx(
            [
                ("documento", "sexo"),  # encabezado: se saltea
                ("30.123.456", "f"),
                (28111222, "MASCULINO"),
                ("", "F"),  # sin dni: rechazada
                ("27000111", "Z"),  # sexo inválido: rechazada
                ("30123456", "M"),  # dni duplicado: rechazada
            ]
        )
        entradas, resumen = parsear_padron(archivo)
        self.assertEqual(_pares(entradas), [("30123456", "F"), ("28111222", "M")])
        self.assertEqual(resumen.rechazadas, 3)
        self.assertEqual(resumen.validas, 2)
        self.assertEqual(resumen.con_identidad, 0)

    def test_sin_encabezado_tambien_funciona(self):
        entradas, resumen = parsear_padron(_xlsx([("30123456", "F")]))
        self.assertEqual(_pares(entradas), [("30123456", "F")])
        self.assertEqual(resumen.rechazadas, 0)

    def test_seis_columnas_con_identidad(self):
        """Cambio 57: nombre, apellido, fecha y localidad viajan por fila; la
        identidad completa se cuenta aparte."""
        entradas, resumen = parsear_padron(
            _xlsx(
                [
                    ("documento", "sexo", "nombre", "apellido", "fecha de nacimiento", "localidad"),
                    ("36210951", "F", " Pamela  Janet ", "Romero", "14/03/2010", "Resistencia"),
                    ("28111222", "M", "", "", "", ""),
                    ("20111333", "F", "Ana", "", "2001-05-09", "Sáenz Peña"),
                ]
            )
        )
        self.assertEqual(resumen.validas, 3)
        self.assertEqual(resumen.con_identidad, 1)  # Ana no tiene apellido
        pamela = entradas[0]
        self.assertEqual(pamela["nombre"], "Pamela Janet")
        self.assertEqual(pamela["apellido"], "Romero")
        self.assertEqual(pamela["fecha_nacimiento"], date(2010, 3, 14))
        self.assertEqual(pamela["localidad_texto"], "Resistencia")
        self.assertEqual(entradas[2]["fecha_nacimiento"], date(2001, 5, 9))

    def test_fecha_invalida_no_rechaza_la_fila(self):
        entradas, resumen = parsear_padron(_xlsx([("30123456", "F", "Ana", "Paz", "ayer", "")]))
        self.assertEqual(len(entradas), 1)
        self.assertIsNone(entradas[0]["fecha_nacimiento"])
        self.assertEqual(resumen.fechas_invalidas, 1)
        self.assertEqual(resumen.con_identidad, 1)

    def test_fecha_como_celda_de_excel(self):
        entradas, _ = parsear_padron(_xlsx([("30123456", "F", "Ana", "Paz", date(1991, 3, 14), "")]))
        self.assertEqual(entradas[0]["fecha_nacimiento"], date(1991, 3, 14))

    def test_extension_invalida(self):
        archivo = SimpleUploadedFile("padron.csv", b"30123456,F", content_type="text/csv")
        with self.assertRaises(ValidationError):
            parsear_padron(archivo)

    def test_contenido_no_excel(self):
        archivo = SimpleUploadedFile("padron.xlsx", b"esto no es un excel", content_type=XLSX_MIME)
        with self.assertRaises(ValidationError):
            parsear_padron(archivo)

    def test_sin_filas_validas(self):
        with self.assertRaises(ValidationError):
            parsear_padron(_xlsx([("documento", "sexo"), ("", "")]))


class NormalizacionTests(TestCase):
    def test_fechas(self):
        self.assertEqual(normalizar_fecha("14/03/2010"), (date(2010, 3, 14), False))
        self.assertEqual(normalizar_fecha("2010-03-14"), (date(2010, 3, 14), False))
        self.assertEqual(normalizar_fecha(None), (None, False))
        self.assertEqual(normalizar_fecha("   "), (None, False))
        self.assertEqual(normalizar_fecha("14/13/2010"), (None, True))
        # Serial de Excel: 40251 = 14/03/2010.
        self.assertEqual(normalizar_fecha(40251), (date(2010, 3, 14), False))

    def test_clave_localidad(self):
        self.assertEqual(clave_localidad("Sáenz Peña"), clave_localidad("SAENZ PENA"))
        self.assertEqual(clave_localidad("  Resistencia "), "resistencia")
        self.assertNotEqual(clave_localidad("Rcia."), clave_localidad("Resistencia"))


class EstaHabilitadoTests(_BasePadronTest):
    def test_sin_padron_el_link_es_abierto(self):
        self.assertTrue(esta_habilitado(self.relevamiento, "99999999", "F"))
        self.assertTrue(esta_habilitado(self.convocatoria, "99999999", "F"))

    def test_matchea_con_normalizacion_en_ambos_sentidos(self):
        cargar_padron(self.relevamiento, None, [("30123456", "F")])
        self.assertTrue(esta_habilitado(self.relevamiento, "30.123.456", "femenino"))
        self.assertFalse(esta_habilitado(self.relevamiento, "30123456", "M"))
        self.assertFalse(esta_habilitado(self.relevamiento, "11111111", "F"))

    def test_reemplazo_total_con_efecto_inmediato(self):
        cargar_padron(self.convocatoria, None, [("30123456", "F")])
        cargar_padron(self.convocatoria, None, [("28111222", "M")])
        self.assertFalse(esta_habilitado(self.relevamiento, "30123456", "F"))
        self.assertTrue(esta_habilitado(self.relevamiento, "28111222", "M"))
        self.assertEqual(self.convocatoria.padron.count(), 1)

    def test_el_padron_es_de_la_convocatoria_y_lo_comparten_sus_relevamientos(self):
        """Cambio 57: un solo Excel para el link y para el territorial."""
        territorial = User.objects.create_user("terri_pad")
        rel_campo = Relevamiento.objects.create(
            convocatoria=self.convocatoria,
            territorial=territorial,
            fecha_asignada=date(2026, 6, 1),
            zona="Zona",
        )
        cargar_padron(self.convocatoria, None, [("30123456", "F")])
        self.assertTrue(esta_habilitado(rel_campo, "30123456", "F"))
        self.assertTrue(esta_habilitado(self.relevamiento, "30123456", "F"))
        self.assertFalse(esta_habilitado(rel_campo, "99999999", "F"))


class CargaConIdentidadTests(_BasePadronTest):
    def setUp(self):
        super().setUp()
        provincia = Provincia.objects.create(nombre="Chaco")
        municipio = Municipio.objects.create(nombre="San Fernando", provincia=provincia)
        self.resistencia = Localidad.objects.create(nombre="Resistencia", municipio=municipio)

    def test_guarda_identidad_y_cruza_localidad_por_nombre(self):
        resumen = cargar_padron(
            self.convocatoria,
            None,
            [
                {
                    "dni": "36210951",
                    "sexo": "F",
                    "nombre": "Pamela Janet",
                    "apellido": "Romero",
                    "fecha_nacimiento": date(2010, 3, 14),
                    "localidad_texto": "RESISTENCIA",
                },
                {"dni": "28111222", "sexo": "M", "localidad_texto": "Rcia."},
            ],
        )
        fila = fila_padron(self.convocatoria, "36.210.951", "femenino")
        self.assertTrue(fila.tiene_identidad)
        self.assertEqual(fila.localidad, self.resistencia)
        self.assertEqual(fila.localidad_texto, "RESISTENCIA")
        otra = fila_padron(self.convocatoria, "28111222", "M")
        self.assertFalse(otra.tiene_identidad)
        self.assertIsNone(otra.localidad)
        self.assertEqual(otra.localidad_texto, "Rcia.")
        self.assertEqual(resumen.validas, 2)
        self.assertEqual(resumen.con_identidad, 1)
        self.assertEqual(resumen.localidades_no_reconocidas, ["Rcia."])

    def test_las_tuplas_historicas_siguen_valiendo(self):
        """RN-7: un padrón de dos columnas habilita y no valida."""
        cargar_padron(self.convocatoria, None, [("30123456", "F")])
        fila = fila_padron(self.convocatoria, "30123456", "F")
        self.assertFalse(fila.tiene_identidad)

    def test_fila_padron_sin_dni_o_sexo_es_none(self):
        cargar_padron(self.convocatoria, None, [("30123456", "F")])
        self.assertIsNone(fila_padron(self.convocatoria, "", "F"))
        self.assertIsNone(fila_padron(self.convocatoria, "30123456", "X"))


class FormSinPadronTests(_BasePadronTest):
    def test_el_alta_de_relevamiento_ya_no_tiene_padron(self):
        form = RelevamientoForm(
            data={
                "tipo": Relevamiento.Tipo.PUBLICO,
                "convocatoria": self.convocatoria.pk,
                "fecha_asignada": "2026-07-01T08:00",
                "fecha_hasta": "2026-07-31T18:00",
            },
            puede_publico=True,
        )
        self.assertNotIn("padron", form.fields)
        self.assertTrue(form.is_valid(), form.errors)
        rel = form.save()
        self.assertTrue(esta_habilitado(rel, "1234567", "F"))


class PadronConvocatoriaViewTests(_BasePadronTest):
    """Alta y reemplazo desde la convocatoria (Cambio 57)."""

    def setUp(self):
        super().setUp()
        call_command("seed_becas", stdout=StringIO())
        grupo_admin = Group.objects.get(name=ROL_ADMIN)
        self.admin = User.objects.create_user("admin_pad", password="x")
        self.admin.groups.add(grupo_admin)
        self.sin_permiso = User.objects.create_user("sin_pad", password="x")

    def _url(self):
        return reverse("becas:convocatoria_padron", args=[self.convocatoria.pk])

    def test_carga_ok_y_redirige_a_la_convocatoria(self):
        self.client.force_login(self.admin)
        resp = self.client.post(
            self._url(),
            {"padron": _xlsx([("documento", "sexo", "nombre", "apellido"), ("30123456", "F", "Ana", "Paz")])},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("becas:convocatoria_detalle", args=[self.convocatoria.pk]), resp.url)
        self.assertTrue(esta_habilitado(self.relevamiento, "30123456", "F"))
        self.convocatoria.refresh_from_db()
        self.assertTrue(self.convocatoria.padron_archivo)

    def test_sin_archivo_avisa(self):
        self.client.force_login(self.admin)
        resp = self.client.post(self._url(), {})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.convocatoria.padron.count(), 0)

    def test_archivo_invalido_no_borra_el_padron_anterior(self):
        cargar_padron(self.convocatoria, None, [("11111111", "M")])
        self.client.force_login(self.admin)
        self.client.post(self._url(), {"padron": SimpleUploadedFile("p.xlsx", b"no excel", content_type=XLSX_MIME)})
        self.assertEqual(list(self.convocatoria.padron.values_list("dni", flat=True)), ["11111111"])

    def test_sin_capacidad_no_puede(self):
        self.client.force_login(self.sin_permiso)
        resp = self.client.post(self._url(), {"padron": _xlsx([("30123456", "F")])})
        self.assertNotEqual(resp.status_code, 200)
        self.assertEqual(self.convocatoria.padron.count(), 0)

    def test_solo_post(self):
        self.client.force_login(self.admin)
        resp = self.client.get(self._url())
        self.assertEqual(resp.status_code, 405)

    def test_la_url_por_relevamiento_es_el_padron_propio(self):
        """Cambio 59: la ruta por relevamiento volvió, pero como padrón PROPIO
        (pisa al de la convocatoria), y solo por POST."""
        url = reverse("becas:relevamiento_padron", args=[self.relevamiento.pk])
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(url).status_code, 405)

    def test_plantilla_descargable(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse("becas:convocatoria_padron_plantilla", args=[self.convocatoria.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], XLSX_MIME)
        self.assertIn("plantilla-padron-habilitados.xlsx", resp["Content-Disposition"])
        from openpyxl import load_workbook

        hoja = load_workbook(BytesIO(resp.content)).active
        encabezados = [c.value for c in next(hoja.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(encabezados, ["documento", "sexo", "nombre", "apellido", "fecha de nacimiento", "localidad"])

    def test_la_plantilla_se_puede_volver_a_cargar(self):
        """El ejemplo que descargamos tiene que pasar por nuestro propio parser."""
        entradas, resumen = parsear_padron(
            SimpleUploadedFile("plantilla.xlsx", plantilla_padron(), content_type=XLSX_MIME)
        )
        self.assertEqual(resumen.validas, 2)
        self.assertEqual(resumen.con_identidad, 1)
        self.assertEqual(entradas[0]["fecha_nacimiento"], date(1991, 3, 14))

    def test_el_detalle_de_la_convocatoria_muestra_el_padron(self):
        cargar_padron(
            self.convocatoria,
            None,
            [{"dni": "30123456", "sexo": "F", "nombre": "Ana", "apellido": "Paz"}, ("28111222", "M")],
        )
        self.client.force_login(self.admin)
        try:
            resp = self.client.get(reverse("becas:convocatoria_detalle", args=[self.convocatoria.pk]))
        except AttributeError as exc:  # bug conocido del test client local (Py3.14 + Dj4.2)
            if "dicts" not in str(exc):
                raise
            return
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Padrón de habilitados")
        self.assertContains(resp, "2 habilitados")
        self.assertContains(resp, reverse("becas:convocatoria_padron_plantilla", args=[self.convocatoria.pk]))


class PadronPorRelevamientoTests(_BasePadronTest):
    """Cambio 59: el padrón de la convocatoria se hereda; el propio de un
    relevamiento lo pisa solo para ese relevamiento."""

    def setUp(self):
        super().setUp()
        territorial = User.objects.create_user("terri_c59")
        self.rel_campo = Relevamiento.objects.create(
            convocatoria=self.convocatoria,
            territorial=territorial,
            fecha_asignada=date(2026, 6, 1),
            zona="Zona",
        )

    def test_hereda_hasta_tener_propio_y_el_propio_pisa(self):
        cargar_padron(self.convocatoria, None, [("30123456", "F")])
        # Los dos relevamientos heredan.
        self.assertTrue(esta_habilitado(self.relevamiento, "30123456", "F"))
        self.assertTrue(esta_habilitado(self.rel_campo, "30123456", "F"))
        # El público carga padrón propio: pisa al heredado SOLO para él.
        cargar_padron(self.relevamiento, None, [("28111222", "M")])
        self.assertFalse(esta_habilitado(self.relevamiento, "30123456", "F"))
        self.assertTrue(esta_habilitado(self.relevamiento, "28111222", "M"))
        self.assertTrue(esta_habilitado(self.rel_campo, "30123456", "F"))  # sigue heredando
        self.assertFalse(esta_habilitado(self.rel_campo, "28111222", "M"))

    def test_quitar_el_propio_vuelve_a_heredar(self):
        from programas.services.padron import origen_padron, quitar_padron_propio

        cargar_padron(self.convocatoria, None, [("30123456", "F")])
        cargar_padron(self.relevamiento, None, [("28111222", "M")])
        self.assertEqual(origen_padron(self.relevamiento), "propio")
        filas = quitar_padron_propio(self.relevamiento)
        self.assertEqual(filas, 1)
        self.assertEqual(origen_padron(self.relevamiento), "convocatoria")
        self.assertTrue(esta_habilitado(self.relevamiento, "30123456", "F"))
        self.assertFalse(esta_habilitado(self.relevamiento, "28111222", "M"))

    def test_identificar_usa_el_padron_efectivo(self):
        from programas.services.identidad import identificar

        cargar_padron(
            self.convocatoria,
            None,
            [{"dni": "30123456", "sexo": "F", "nombre": "Ana", "apellido": "Paz"}],
        )
        cargar_padron(
            self.relevamiento,
            None,
            [{"dni": "30123456", "sexo": "F", "nombre": "Maria", "apellido": "Gomez"}],
        )
        con_propio = identificar(self.relevamiento, "30123456", "F")
        heredado = identificar(self.rel_campo, "30123456", "F")
        self.assertEqual(con_propio["datos"]["nombre"], "Maria")
        self.assertEqual(heredado["datos"]["nombre"], "Ana")

    def test_la_carga_valida_los_casos_de_su_alcance(self):
        from programas.models import Formulario

        pendiente_publico = Formulario.objects.create(
            relevamiento=self.relevamiento, datos_identificacion={"dni": "30123456", "sexo": "F"}
        )
        pendiente_campo = Formulario.objects.create(
            relevamiento=self.rel_campo, datos_identificacion={"dni": "30123456", "sexo": "F"}
        )
        # El público tiene padrón propio SIN identidad: su caso no valida acá.
        cargar_padron(self.relevamiento, None, [("30123456", "F")])
        resumen = cargar_padron(
            self.convocatoria,
            None,
            [{"dni": "30123456", "sexo": "F", "nombre": "Ana", "apellido": "Paz"}],
        )
        pendiente_publico.refresh_from_db()
        pendiente_campo.refresh_from_db()
        self.assertEqual(resumen.casos_validados, 1)
        self.assertTrue(pendiente_campo.validado_renaper)  # hereda: lo valida la convocatoria
        self.assertFalse(pendiente_publico.validado_renaper)  # su padrón propio manda
        # Reemplazo el propio por uno con identidad: ahora valida su caso.
        resumen = cargar_padron(
            self.relevamiento,
            None,
            [{"dni": "30123456", "sexo": "F", "nombre": "Maria", "apellido": "Gomez"}],
        )
        pendiente_publico.refresh_from_db()
        self.assertEqual(resumen.casos_validados, 1)
        self.assertTrue(pendiente_publico.validado_renaper)

    def test_objetivo_con_identidad_prefiere_el_efectivo(self):
        from programas.services.padron import objetivo_con_identidad

        cargar_padron(
            self.convocatoria,
            None,
            [{"dni": "30123456", "sexo": "F", "nombre": "Ana", "apellido": "Paz"}],
        )
        # Sin propios: cualquiera de los dos sirve (el primero de la lista).
        elegido = objetivo_con_identidad([self.rel_campo, self.relevamiento], "30123456", "F")
        self.assertEqual(elegido, self.rel_campo)
        # El de campo carga un propio SIN esa persona: deja de servir.
        cargar_padron(self.rel_campo, None, [("28111222", "M")])
        elegido = objetivo_con_identidad([self.rel_campo, self.relevamiento], "30123456", "F")
        self.assertEqual(elegido, self.relevamiento)


class PadronRelevamientoViewTests(_BasePadronTest):
    """Carga y quita del padrón propio desde el detalle del relevamiento."""

    def setUp(self):
        super().setUp()
        call_command("seed_becas", stdout=StringIO())
        self.admin = User.objects.create_user("admin_pad59", password="x")
        grupo = Group.objects.get(name=ROL_ADMIN)
        # El relevamiento del fixture es público: el alcance exige la capacidad
        # de públicos (en la base real la asigna users.0025; acá, syncdb).
        from django.contrib.auth.models import Permission

        from core.rbac import APP_LABEL, codename_de

        grupo.permissions.add(
            Permission.objects.get(
                content_type__app_label=APP_LABEL, codename=codename_de("becas.relevamiento.publico")
            )
        )
        self.admin.groups.add(grupo)

    def test_carga_quita_y_permisos(self):
        self.client.force_login(self.admin)
        url = reverse("becas:relevamiento_padron", args=[self.relevamiento.pk])
        resp = self.client.post(
            url, {"padron": _xlsx([("documento", "sexo", "nombre", "apellido"), ("30123456", "F", "Ana", "Paz")])}
        )
        self.assertEqual(resp.status_code, 302)
        self.assertIn(reverse("becas:relevamiento_detalle", args=[self.relevamiento.pk]), resp.url)
        self.relevamiento.refresh_from_db()
        self.assertEqual(self.relevamiento.padron_propio.count(), 1)
        self.assertTrue(self.relevamiento.padron_archivo)

        resp = self.client.post(reverse("becas:relevamiento_padron_quitar", args=[self.relevamiento.pk]))
        self.assertEqual(resp.status_code, 302)
        self.relevamiento.refresh_from_db()
        self.assertEqual(self.relevamiento.padron_propio.count(), 0)
        self.assertFalse(self.relevamiento.padron_archivo)

        sin_permiso = User.objects.create_user("sin_pad59", password="x")
        self.client.force_login(sin_permiso)
        resp = self.client.post(url, {"padron": _xlsx([("30123456", "F")])})
        self.assertNotEqual(resp.status_code, 200)
        self.assertEqual(self.relevamiento.padron_propio.count(), 0)
