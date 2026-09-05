"""Dashboard del programa Becas (análisis #366): servicio, respuestas, endpoint y exportación."""

import csv
from datetime import date, datetime, time, timedelta
from io import BytesIO, StringIO

from django.apps import apps
from django.contrib.auth.models import Group, Permission, User
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache
from django.core.management import call_command
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core import rbac
from legajos.models import Ciudadano
from programas.forms_reportes import DashboardBecasFiltroForm
from programas.management.commands.seed_becas import ROL_ADMIN, ROL_COORDINADOR_REGIONAL, ROL_TERRITORIAL
from programas.models import (
    AsignacionCoordinador,
    Convocatoria,
    Formulario,
    ListaEspera,
    PreguntaGlobal,
    ProgramaSiis,
    Relevamiento,
    RequisitoNativo,
    Segmento,
    Subsegmento,
    TipoCampo,
    ValidacionSIS,
)
from programas.services import dashboard_becas as svc
from programas.services.autorizacion import programa_becas
from users.models import Capacidad, RolMeta

HOY = date(2026, 9, 5)
VENTANA = svc.Filtros(desde=HOY - timedelta(days=89), hasta=HOY)  # «últimos 90 días»


class DashboardBecasBase(TestCase):
    def setUp(self):
        cache.clear()
        call_command("seed_becas", stdout=StringIO())
        self.programa = ProgramaSiis.objects.create(
            nombre="Incentivo Juventud", siis_programa_id=901, siis_programa_estado=ProgramaSiis.EstadoSiis.ACTIVO
        )
        self.otro_programa = ProgramaSiis.objects.create(
            nombre="Otro programa", siis_programa_id=902, siis_programa_estado=ProgramaSiis.EstadoSiis.ACTIVO
        )
        self.segmento = Segmento.objects.create(programa=self.programa, nombre="Secundario", cupo_maximo=300)
        self.segmento_b = Segmento.objects.create(programa=self.programa, nombre="Chaco Olímpico", cupo_maximo=50)
        self.segmento_otro = Segmento.objects.create(
            programa=self.otro_programa, nombre="Ajeno al programa", cupo_maximo=10
        )

        self.admin = self._usuario("admin-dash", ROL_ADMIN)
        self.regional = self._usuario("regional-dash", ROL_COORDINADOR_REGIONAL)
        self.otro_regional = self._usuario("regional-ajeno-dash", ROL_COORDINADOR_REGIONAL)
        self.territorial = self._usuario("territorial-dash", ROL_TERRITORIAL, first_name="Marta", last_name="Gómez")
        self.territorial_b = self._usuario("territorial-b-dash", ROL_TERRITORIAL, first_name="Luis", last_name="Ojeda")

        self.sub_propio = Subsegmento.objects.create(
            segmento=self.segmento, nombre="Resistencia", cupo_maximo=100, referente=self.regional
        )
        self.sub_ajeno = Subsegmento.objects.create(
            segmento=self.segmento, nombre="Interior", cupo_maximo=150, referente=self.otro_regional
        )
        self.conv_propia = self._convocatoria("Secundario Resistencia 2026", self.segmento, self.sub_propio)
        self.conv_ajena = self._convocatoria("Secundario Interior 2026", self.segmento, self.sub_ajeno)
        self.conv_b = self._convocatoria(
            "Chaco Olímpico 2026", self.segmento_b, activo=False, cerrada_automaticamente=True
        )
        self.conv_otro = self._convocatoria("Convocatoria de otro programa", self.segmento_otro)

        self.rel_propio = self._relevamiento(self.conv_propia, self.territorial, Relevamiento.Estado.TERMINADO)
        self.rel_publico = self._relevamiento(self.conv_propia, None, Relevamiento.Estado.EN_CURSO, publico=True)
        self.rel_ajeno = self._relevamiento(self.conv_ajena, self.territorial_b, Relevamiento.Estado.EN_REVISION)
        self.rel_b = self._relevamiento(self.conv_b, self.territorial_b, Relevamiento.Estado.ASIGNADO)
        self.rel_otro = self._relevamiento(self.conv_otro, self.territorial, Relevamiento.Estado.TERMINADO)

        self.q_laboral = PreguntaGlobal.objects.create(
            texto="Situación laboral",
            tipo=TipoCampo.SELECTOR,
            opciones=["Trabaja", "Estudia", "Busca trabajo"],
            orden=1,
        )
        self.q_transporte = PreguntaGlobal.objects.create(
            texto="¿Cómo llegás?", tipo=TipoCampo.SELECTOR_MULTIPLE, opciones=["Colectivo", "A pie", "Moto"], orden=2
        )
        PreguntaGlobal.objects.create(texto="Observaciones", tipo=TipoCampo.STRING, orden=3)
        PreguntaGlobal.objects.create(texto="Inactiva", tipo=TipoCampo.SELECTOR, opciones=["A", "B"], activo=False)
        self.r_cursa = RequisitoNativo.objects.create(
            programa=self.programa, texto="¿Cursás actualmente?", tipo=TipoCampo.SELECTOR, opciones=["Sí", "No"]
        )
        RequisitoNativo.objects.create(
            programa=self.otro_programa, texto="Requisito ajeno", tipo=TipoCampo.SELECTOR, opciones=["X"]
        )

    # --- helpers -----------------------------------------------------------
    def _usuario(self, username, rol, **extra):
        usuario = User.objects.create_user(username, password="x", **extra)
        usuario.groups.add(Group.objects.get(name=rol))
        return usuario

    def _convocatoria(self, nombre, segmento, subsegmento=None, **extra):
        valores = {
            "nombre": nombre,
            "segmento": segmento,
            "subsegmento": subsegmento,
            "fecha_inicio": date(2026, 3, 1),
            "fecha_fin": date(2026, 11, 30),
        }
        valores.update(extra)
        return Convocatoria.objects.create(**valores)

    def _relevamiento(self, convocatoria, territorial, estado, publico=False):
        return Relevamiento.objects.create(
            convocatoria=convocatoria,
            tipo=Relevamiento.Tipo.PUBLICO if publico else Relevamiento.Tipo.TERRITORIAL,
            territorial=territorial,
            fecha_asignada=date(2026, 6, 1),
            fecha_hasta=date(2026, 11, 30),
            zona="" if publico else "Zona",
            estado=estado,
        )

    def _formulario(self, relevamiento, estado=Formulario.Estado.ENVIADO, creado=None, **extra):
        valores = {
            "relevamiento": relevamiento,
            "celular": "3624000000",
            "email_contacto": "dash@example.com",
            "estado": estado,
            "datos_identificacion": {"dni": str(40000000 + Formulario.objects.count()), "nombre": "P", "apellido": "P"},
        }
        valores.update(extra)
        formulario = Formulario.objects.create(**valores)
        if creado is not None:
            momento = timezone.make_aware(datetime.combine(creado, time(12, 0)), timezone.get_current_timezone())
            Formulario.objects.filter(pk=formulario.pk).update(creado=momento)
            formulario.refresh_from_db()
        return formulario

    def _ciudadano(self, dni, localidad=None):
        return Ciudadano.objects.create(dni=dni, nombre="Ana", apellido="Pérez", localidad=localidad)

    def _usuario_sin_reportes(self):
        """Ve el programa (coordina el segmento) pero no tiene ``becas.reportes.ver``."""
        grupo = Group.objects.create(name="Becas — Solo segmentos")
        RolMeta.objects.create(grupo=grupo, programa=programa_becas(), activo=True)
        content_type = ContentType.objects.get_for_model(Capacidad)
        permiso = Permission.objects.get(content_type=content_type, codename=rbac.codename_de("becas.segmento.ver"))
        grupo.permissions.add(permiso)
        usuario = User.objects.create_user("solo-segmentos", password="x")
        usuario.groups.add(grupo)
        AsignacionCoordinador.objects.create(segmento=self.segmento, coordinador=usuario)
        return usuario


class MetricasTests(DashboardBecasBase):
    def test_totales_cierran_entre_bloques(self):
        dentro = HOY - timedelta(days=10)
        for estado in (Formulario.Estado.ENVIADO, Formulario.Estado.ENVIADO, Formulario.Estado.APROBADO):
            self._formulario(self.rel_propio, estado, creado=dentro)
        self._formulario(self.rel_publico, Formulario.Estado.RECHAZADO, creado=dentro - timedelta(days=21))
        self._formulario(self.rel_ajeno, Formulario.Estado.BAJA, creado=dentro)
        self._formulario(self.rel_propio, Formulario.Estado.APROBADO, creado=HOY - timedelta(days=200))  # fuera
        self._formulario(self.rel_otro, Formulario.Estado.APROBADO, creado=dentro)  # otro programa

        datos = svc.metricas(self.admin, self.programa, VENTANA)

        i = datos.indicadores
        self.assertEqual(i.formularios_recibidos, 5)
        self.assertEqual(sum(f["total"] for f in datos.serie_semanal), 5)
        self.assertEqual(sum(f["total"] for f in datos.estados), 5)
        self.assertEqual(sum(c["recibidos"] for c in datos.convocatorias), 5)
        self.assertEqual(sum(c["total"] for c in datos.canales), 5)
        self.assertEqual(i.aprobados, 1)
        self.assertEqual(i.pendientes, 2)
        self.assertEqual(i.tasa_aprobacion, 20.0)
        self.assertEqual({f["clave"]: f["total"] for f in datos.estados}["RECHAZADO"], 1)
        self.assertEqual(datos.embudo[0], {"etapa": "Formularios recibidos", "total": 5, "pct": 100.0})
        # La serie no salta semanas: cubre la ventana completa (13 semanas de 90 días).
        self.assertGreaterEqual(len(datos.serie_semanal), 13)
        self.assertIn("Del 08/06/2026 al 05/09/2026", datos.alcance)

    def test_convocatorias_y_relevamientos_cuentan_por_estructura(self):
        datos = svc.metricas(self.admin, self.programa, svc.Filtros(desde=HOY, hasta=HOY))

        i = datos.indicadores
        self.assertEqual(i.convocatorias_total, 3)
        self.assertEqual(i.convocatorias_activas, 2)
        self.assertEqual(i.convocatorias_cerradas_vencimiento, 1)
        self.assertEqual(i.relevamientos_total, 4)  # el de otro programa no entra
        self.assertEqual(i.relevamientos_en_curso, 1)
        self.assertEqual(i.relevamientos_publicos, 1)
        por_estado = {f["clave"]: f["total"] for f in datos.relevamientos_por_estado}
        self.assertEqual(por_estado["ASIGNADO"], 1)
        self.assertEqual(len(por_estado), len(Relevamiento.Estado.choices))
        nombres = [c["nombre"] for c in datos.convocatorias]
        self.assertIn("Chaco Olímpico 2026", nombres)
        self.assertNotIn("Convocatoria de otro programa", nombres)
        cerrada = next(c for c in datos.convocatorias if c["nombre"] == "Chaco Olímpico 2026")
        self.assertEqual(cerrada["estado"], "Cerrada por vencimiento")

    def test_regional_solo_ve_su_subsegmento(self):
        dentro = HOY - timedelta(days=5)
        self._formulario(self.rel_propio, Formulario.Estado.APROBADO, creado=dentro)
        self._formulario(self.rel_ajeno, Formulario.Estado.APROBADO, creado=dentro)
        self._formulario(self.rel_ajeno, Formulario.Estado.APROBADO, creado=dentro)

        datos = svc.metricas(self.regional, self.programa, VENTANA)

        self.assertEqual(datos.indicadores.formularios_recibidos, 1)
        self.assertEqual([c["nombre"] for c in datos.convocatorias], ["Secundario Resistencia 2026"])
        self.assertEqual(datos.indicadores.cupo_total, 100)  # distribuido en su subsegmento, no los 300 del segmento
        self.assertEqual(datos.indicadores.cupo_ocupado, 1)
        self.assertNotIn("Interior", str(datos.to_dict()))
        self.assertEqual([t["nombre"] for t in datos.territoriales], ["Marta Gómez"])

    def test_cupo_no_depende_del_periodo(self):
        self._formulario(self.rel_propio, Formulario.Estado.APROBADO, creado=HOY - timedelta(days=300))
        self._formulario(self.rel_propio, Formulario.Estado.APROBADO, creado=HOY - timedelta(days=2))

        datos = svc.metricas(self.admin, self.programa, svc.Filtros(desde=HOY - timedelta(days=6), hasta=HOY))

        self.assertEqual(datos.indicadores.formularios_recibidos, 1)
        self.assertEqual(datos.indicadores.cupo_ocupado, 2)
        self.assertEqual(datos.indicadores.cupo_total, 350)
        fila = next(c for c in datos.convocatorias if c["id"] == self.conv_propia.pk)
        self.assertEqual((fila["cupo_segmento"], fila["cupo_ocupado"]), (300, 2))

    def test_variacion_contra_periodo_anterior(self):
        ventana = svc.Filtros(desde=HOY - timedelta(days=29), hasta=HOY)
        for _ in range(3):
            self._formulario(self.rel_propio, creado=HOY - timedelta(days=3))
        for _ in range(2):
            self._formulario(self.rel_propio, creado=HOY - timedelta(days=40))

        self.assertEqual(svc.metricas(self.admin, self.programa, ventana).indicadores.variacion_periodo_anterior, 50)
        self.assertIsNone(svc.metricas(self.admin, self.programa, svc.Filtros()).indicadores.variacion_periodo_anterior)
        Formulario.objects.filter(creado__lt=timezone.now() - timedelta(days=30)).delete()
        self.assertIsNone(svc.metricas(self.admin, self.programa, ventana).indicadores.variacion_periodo_anterior)

    def test_embudo_lista_de_espera_y_canales(self):
        dentro = HOY - timedelta(days=1)
        validado = self._formulario(self.rel_propio, Formulario.Estado.APROBADO, creado=dentro, validado_renaper=True)
        forzado = self._formulario(self.rel_publico, Formulario.Estado.APROBADO, creado=dentro, identidad_forzada=True)
        espera = self._formulario(self.rel_publico, Formulario.Estado.ENVIADO, creado=dentro)
        self._formulario(self.rel_propio, Formulario.Estado.RECHAZADO, creado=dentro)
        ValidacionSIS.objects.create(formulario=validado, estado=ValidacionSIS.Estado.RECHAZADO, documento="1")
        ValidacionSIS.objects.create(
            formulario=validado, estado=ValidacionSIS.Estado.OK, documento="1"
        )  # la última manda
        ValidacionSIS.objects.create(formulario=forzado, estado=ValidacionSIS.Estado.RECHAZADO, documento="2")
        ListaEspera.objects.create(formulario=espera, segmento=self.segmento, posicion=1)
        ListaEspera.objects.create(
            formulario=self._formulario(self.rel_propio, creado=dentro),
            segmento=self.segmento,
            posicion=2,
            promovido=True,
        )

        datos = svc.metricas(self.admin, self.programa, VENTANA)

        embudo = {f["etapa"]: f["total"] for f in datos.embudo}
        self.assertEqual(embudo["Formularios recibidos"], 5)
        self.assertEqual(embudo["Identidad validada"], 2)
        self.assertEqual(embudo["Aprobados"], 2)
        self.assertEqual(embudo["Validación SIIS OK"], 1)
        self.assertEqual(embudo["En lista de espera"], 1)
        self.assertEqual(embudo["Rechazados"], 1)
        self.assertEqual(datos.indicadores.lista_espera, 1)
        canales = {c["clave"]: c["total"] for c in datos.canales}
        self.assertEqual(canales, {"TERRITORIAL": 3, "PUBLICO": 2})

    def test_filtros_de_canal_y_relevamiento(self):
        dentro = HOY - timedelta(days=1)
        self._formulario(self.rel_propio, creado=dentro)
        self._formulario(self.rel_publico, creado=dentro)
        self._formulario(self.rel_publico, creado=dentro)

        publico = svc.metricas(self.admin, self.programa, svc.Filtros(canal=Relevamiento.Tipo.PUBLICO))
        self.assertEqual(publico.indicadores.formularios_recibidos, 2)
        self.assertEqual(publico.indicadores.relevamientos_total, 1)
        self.assertEqual(publico.territoriales, [])
        self.assertIn("Link público", publico.alcance)

        uno = svc.metricas(self.admin, self.programa, svc.Filtros(relevamiento_id=self.rel_propio.pk))
        self.assertEqual(uno.indicadores.formularios_recibidos, 1)
        self.assertEqual(uno.indicadores.convocatorias_total, 1)
        self.assertIn(self.rel_propio.nombre, uno.alcance)

    def test_territoriales_y_localidades(self):
        Localidad = apps.get_model("core", "Localidad")
        resistencia = Localidad.objects.create(nombre="Resistencia")
        dentro = HOY - timedelta(days=1)
        self._formulario(
            self.rel_propio, Formulario.Estado.APROBADO, creado=dentro, ciudadano=self._ciudadano("1", resistencia)
        )
        self._formulario(self.rel_propio, creado=dentro, ciudadano=self._ciudadano("2", resistencia))
        self._formulario(self.rel_publico, creado=dentro, ciudadano=self._ciudadano("3"))
        self._formulario(self.rel_ajeno, creado=dentro)

        datos = svc.metricas(self.admin, self.programa, VENTANA)

        self.assertEqual(
            datos.territoriales[:2],
            [
                {"nombre": "Marta Gómez", "formularios": 2, "aprobados": 1, "relevamientos": 1},
                {"nombre": "Luis Ojeda", "formularios": 1, "aprobados": 0, "relevamientos": 2},
            ],
        )
        self.assertEqual(datos.localidades["top"][0], {"localidad": "Resistencia", "total": 2, "pct": 50.0})
        self.assertEqual(datos.localidades["top"][1]["localidad"], svc.SIN_LOCALIDAD)

    def test_sin_datos_no_rompe(self):
        datos = svc.metricas(self.admin, self.otro_programa, VENTANA)
        self.assertEqual(datos.indicadores.formularios_recibidos, 0)
        self.assertEqual(datos.embudo[0]["pct"], 0.0)
        self.assertEqual(datos.localidades, {"top": [], "detalle": []})
        vacio = svc.metricas(self.admin, self.otro_programa, svc.Filtros())
        self.assertEqual(vacio.serie_semanal, [])

    def test_presupuesto_de_consultas_no_crece_con_los_formularios(self):
        svc.metricas(self.admin, self.programa, VENTANA)
        with CaptureQueriesContext(connection) as pocas:
            svc.metricas(self.admin, self.programa, VENTANA)
        for _ in range(25):
            self._formulario(self.rel_propio, creado=HOY - timedelta(days=3))
            self._formulario(self.rel_publico, Formulario.Estado.APROBADO, creado=HOY - timedelta(days=9))
        with CaptureQueriesContext(connection) as muchas:
            svc.metricas(self.admin, self.programa, VENTANA)
        self.assertEqual(len(muchas), len(pocas))


class RespuestasTests(DashboardBecasBase):
    def _con_respuestas(self):
        dentro = HOY - timedelta(days=2)
        q1, q2, r = str(self.q_laboral.pk), str(self.q_transporte.pk), str(self.r_cursa.pk)
        self._formulario(
            self.rel_propio,
            creado=dentro,
            data={"globales": {q1: "Trabaja", q2: ["Colectivo", "Moto"]}, "requisitos": {r: "Sí"}},
        )
        self._formulario(
            self.rel_propio,
            creado=dentro,
            data={"globales": {q1: "Estudia", q2: ["Colectivo"]}, "requisitos": {r: "Sí"}},
        )
        self._formulario(
            self.rel_publico, creado=dentro, data={"globales": {q1: "Trabaja", q2: ["Bicicleta"]}, "requisitos": {}}
        )
        self._formulario(
            self.rel_publico, creado=dentro, data={"globales": {}, "requisitos": {}}
        )  # anterior al alta: sin la pregunta
        self._formulario(
            self.rel_ajeno, creado=dentro, data={"globales": {q1: "Busca trabajo"}, "requisitos": {r: "No"}}
        )

    def test_catalogo_solo_opciones_cerradas_con_origen(self):
        preguntas = svc.preguntas_graficables(self.admin, self.programa)
        claves = {p.clave: p for p in preguntas}
        self.assertEqual(
            set(claves),
            {f"global:{self.q_laboral.pk}", f"global:{self.q_transporte.pk}", f"requisito:{self.r_cursa.pk}"},
        )
        self.assertEqual(claves[f"requisito:{self.r_cursa.pk}"].origen, "Requisito del programa")
        self.assertTrue(claves[f"global:{self.q_transporte.pk}"].multiple)
        self.assertFalse(claves[f"global:{self.q_laboral.pk}"].multiple)

    def test_lectura_unica_de_la_respuesta(self):
        data = {"globales": {"7": "Trabaja", "8": ["A", "", None, "B"]}, "requisitos": {"3": 12}}
        self.assertEqual(svc.respuesta_de(data, "global:7"), ["Trabaja"])
        self.assertEqual(svc.respuesta_de(data, "global:8"), ["A", "B"])
        self.assertEqual(svc.respuesta_de(data, "requisito:3"), ["12"])
        self.assertEqual(svc.respuesta_de(data, "global:99"), [])
        self.assertEqual(svc.respuesta_de(None, "global:7"), [])

    def test_selector_simple_suma_la_base_y_multiple_puede_superarla(self):
        self._con_respuestas()

        laboral = svc.distribucion_respuestas(self.admin, self.programa, VENTANA, f"global:{self.q_laboral.pk}")
        self.assertEqual(laboral.base, 4)
        self.assertFalse(laboral.multiple)
        self.assertEqual(sum(o["total"] for o in laboral.opciones), laboral.base)
        self.assertEqual(laboral.opciones[0], {"opcion": "Trabaja", "total": 2, "pct": 50.0})

        transporte = svc.distribucion_respuestas(self.admin, self.programa, VENTANA, f"global:{self.q_transporte.pk}")
        self.assertEqual(transporte.base, 3)
        self.assertTrue(transporte.multiple)
        self.assertGreater(sum(o["total"] for o in transporte.opciones), transporte.base)
        etiquetas = [o["opcion"] for o in transporte.opciones]
        self.assertIn("Bicicleta", etiquetas)  # respondida y ya no en el catálogo
        self.assertIn("A pie", etiquetas)  # en el catálogo con 0
        self.assertEqual(transporte.opciones[0], {"opcion": "Colectivo", "total": 2, "pct": 66.7})

    def test_respuestas_respetan_el_alcance(self):
        self._con_respuestas()
        laboral = svc.distribucion_respuestas(self.regional, self.programa, VENTANA, f"global:{self.q_laboral.pk}")
        self.assertEqual(laboral.base, 3)
        self.assertEqual(next(o for o in laboral.opciones if o["opcion"] == "Busca trabajo")["total"], 0)
        with self.assertRaises(ValueError):
            svc.distribucion_respuestas(self.admin, self.programa, VENTANA, "global:999999")

    def test_todas_las_distribuciones_en_una_pasada(self):
        self._con_respuestas()
        svc.distribuciones_respuestas(self.admin, self.programa, VENTANA)
        with CaptureQueriesContext(connection) as consultas:
            todas = svc.distribuciones_respuestas(self.admin, self.programa, VENTANA)
        self.assertEqual(len(todas), 3)
        self.assertLessEqual(len(consultas), 12)


class CacheTests(DashboardBecasBase):
    def test_cache_por_filtros_y_alcance(self):
        self._formulario(self.rel_propio, creado=HOY - timedelta(days=1))
        primero, desde_cache = svc.metricas_cacheadas(self.admin, self.programa, VENTANA)
        self.assertFalse(desde_cache)
        self._formulario(self.rel_propio, creado=HOY - timedelta(days=1))
        segundo, desde_cache = svc.metricas_cacheadas(self.admin, self.programa, VENTANA)
        self.assertTrue(desde_cache)
        self.assertEqual(segundo.indicadores.formularios_recibidos, primero.indicadores.formularios_recibidos)
        tercero, _ = svc.metricas_cacheadas(self.admin, self.programa, VENTANA, recalcular=True)
        self.assertEqual(tercero.indicadores.formularios_recibidos, 2)
        # RN-18: el regional nunca comparte entrada con el admin.
        self.assertNotEqual(
            svc.clave_cache(self.admin, self.programa, VENTANA), svc.clave_cache(self.regional, self.programa, VENTANA)
        )
        regional, desde_cache = svc.metricas_cacheadas(self.regional, self.programa, VENTANA)
        self.assertFalse(desde_cache)
        self.assertNotEqual(
            svc.clave_cache(self.admin, self.programa, VENTANA),
            svc.clave_cache(self.admin, self.programa, svc.Filtros()),
        )


class FormTests(DashboardBecasBase):
    def test_periodo_a_fechas_y_limpieza_dependiente(self):
        form = DashboardBecasFiltroForm(
            {
                "periodo": "30",
                "segmento": self.segmento_b.pk,
                "convocatoria": self.conv_propia.pk,
                "relevamiento": self.rel_propio.pk,
            },
            user=self.admin,
            programa=self.programa,
        )
        self.assertTrue(form.is_valid(), form.errors)
        filtros = form.filtros()
        self.assertEqual((filtros.hasta - filtros.desde).days, 29)
        self.assertEqual(filtros.segmento_id, self.segmento_b.pk)
        self.assertIsNone(filtros.convocatoria_id)  # RN-6: de otro segmento, se limpia
        self.assertIsNone(filtros.relevamiento_id)  # RN-5: sin convocatoria, se limpia
        self.assertEqual(form.clave_pregunta(), f"global:{self.q_laboral.pk}")

    def test_periodo_personalizado_invalido_y_todo(self):
        form = DashboardBecasFiltroForm(
            {"periodo": "custom", "desde": "2026-09-05", "hasta": "2026-09-01"}, user=self.admin, programa=self.programa
        )
        self.assertFalse(form.is_valid())
        form = DashboardBecasFiltroForm(
            {"periodo": "todo", "pregunta": "global:404"}, user=self.admin, programa=self.programa
        )
        self.assertTrue(form.is_valid(), form.errors)
        self.assertFalse(form.filtros().con_ventana)
        self.assertEqual(form.clave_pregunta(), f"global:{self.q_laboral.pk}")  # clave vieja cae a la primera

    def test_querysets_recortados_por_programa_y_alcance(self):
        form = DashboardBecasFiltroForm(user=self.regional, programa=self.programa)
        self.assertEqual(list(form.fields["convocatoria"].queryset), [self.conv_propia])
        self.assertNotIn(self.segmento_otro, form.fields["segmento"].queryset)
        opciones = form.relevamientos_de(self.conv_propia)
        self.assertEqual({o["id"] for o in opciones}, {self.rel_propio.pk, self.rel_publico.pk})


class EndpointTests(DashboardBecasBase):
    def _url(self, **params):
        base = reverse("becas:programa_dashboard_datos", args=[self.programa.pk])
        if params:
            base += "?" + "&".join(f"{k}={v}" for k, v in params.items())
        return base

    def test_admin_recibe_datos_coherentes_con_el_servicio(self):
        self._formulario(
            self.rel_propio,
            creado=HOY - timedelta(days=1),
            data={"globales": {str(self.q_laboral.pk): "Trabaja"}, "requisitos": {}},
        )
        self.client.force_login(self.admin)

        respuesta = self.client.get(self._url(periodo="90", convocatoria=self.conv_propia.pk))

        self.assertEqual(respuesta.status_code, 200)
        cuerpo = respuesta.json()
        self.assertEqual(set(cuerpo), {"datos", "desde_cache", "respuestas", "opciones", "filtros_aplicados", "avisos"})
        self.assertEqual(cuerpo["datos"]["indicadores"]["formularios_recibidos"], 1)
        self.assertEqual(cuerpo["respuestas"]["clave"], f"global:{self.q_laboral.pk}")
        self.assertEqual(len(cuerpo["opciones"]["relevamientos"]), 2)
        self.assertEqual(cuerpo["filtros_aplicados"]["convocatoria"], self.conv_propia.pk)
        self.assertFalse(cuerpo["desde_cache"])
        self.assertTrue(
            self.client.get(self._url(periodo="90", convocatoria=self.conv_propia.pk)).json()["desde_cache"]
        )

    def test_filtros_invalidos_devuelven_400(self):
        self.client.force_login(self.admin)
        respuesta = self.client.get(self._url(periodo="custom", desde="2026-09-05", hasta="2026-01-01"))
        self.assertEqual(respuesta.status_code, 400)
        self.assertIn("errores", respuesta.json())

    def test_sin_capacidad_de_reportes_no_ve_la_solapa_y_el_endpoint_da_403(self):
        usuario = self._usuario_sin_reportes()
        self.client.force_login(usuario)
        pantalla = self.client.get(reverse("becas:programa_detalle", args=[self.programa.pk]))
        self.assertEqual(pantalla.status_code, 200)
        self.assertNotContains(pantalla, 'id="tab-dashboard"')
        self.assertEqual(self.client.get(self._url()).status_code, 403)

    def test_admin_ve_la_solapa_y_el_regional_no_ve_programas_ajenos(self):
        self.client.force_login(self.admin)
        pantalla = self.client.get(reverse("becas:programa_detalle", args=[self.programa.pk]))
        self.assertContains(pantalla, 'id="tab-dashboard"')
        self.assertContains(pantalla, reverse("becas:programa_dashboard_datos", args=[self.programa.pk]))
        self.client.force_login(self.regional)
        ajeno = self.client.get(reverse("becas:programa_dashboard_datos", args=[self.otro_programa.pk]))
        self.assertEqual(ajeno.status_code, 403)
        propio = self.client.get(self._url())
        self.assertEqual(propio.status_code, 200)
        self.assertEqual(propio.json()["datos"]["indicadores"]["cupo_total"], 100)

    def test_anonimo_redirige_al_login(self):
        respuesta = self.client.get(self._url())
        self.assertEqual(respuesta.status_code, 302)


class ExportacionTests(DashboardBecasBase):
    def _url(self, formato, **params):
        base = reverse("becas:programa_dashboard_exportar", args=[self.programa.pk, formato])
        if params:
            base += "?" + "&".join(f"{k}={v}" for k, v in params.items())
        return base

    def test_xlsx_una_hoja_por_bloque_con_alcance(self):
        self._formulario(
            self.rel_propio,
            Formulario.Estado.APROBADO,
            creado=HOY - timedelta(days=1),
            data={"globales": {str(self.q_laboral.pk): "=Trabaja"}, "requisitos": {}},
        )
        self.client.force_login(self.admin)

        respuesta = self.client.get(self._url("xlsx", periodo="90"))

        self.assertEqual(respuesta.status_code, 200)
        self.assertIn("becas_dashboard_incentivo-juventud_", respuesta["Content-Disposition"])
        libro = load_workbook(BytesIO(respuesta.content), read_only=True)
        self.assertEqual(
            libro.sheetnames,
            [
                "Resumen",
                "Semanas",
                "Estados",
                "Canales",
                "Convocatorias",
                "Relevamientos",
                "Embudo",
                "Territoriales",
                "Localidades",
                "Respuestas",
            ],
        )
        filas = list(libro["Convocatorias"].iter_rows(values_only=True))
        self.assertTrue(str(filas[0][0]).startswith("Alcance: Del "))
        self.assertEqual(filas[2][0], "Convocatoria")
        self.assertIn("Secundario Resistencia 2026", [f[0] for f in filas[3:]])
        resumen = {f[0]: f[1] for f in list(libro["Resumen"].iter_rows(values_only=True))[2:]}
        self.assertEqual(resumen["Formularios recibidos"], 1)
        respuestas = list(libro["Respuestas"].iter_rows(values_only=True))
        self.assertIn("'=Trabaja", [f[3] for f in respuestas[3:]])  # fórmula neutralizada

    def test_csv_de_un_bloque_y_errores(self):
        self._formulario(self.rel_propio, creado=HOY - timedelta(days=1))
        self.client.force_login(self.admin)

        respuesta = self.client.get(self._url("csv", periodo="90", bloque="convocatorias"))
        self.assertEqual(respuesta.status_code, 200)
        contenido = respuesta.content.decode("utf-8-sig")
        filas = list(csv.reader(StringIO(contenido)))
        self.assertTrue(filas[0][0].startswith("Alcance: "))
        self.assertEqual(filas[2][0], "Convocatoria")
        self.assertIn("Secundario Resistencia 2026", [f[0] for f in filas[3:]])

        self.assertEqual(self.client.get(self._url("csv", bloque="inexistente")).status_code, 400)
        self.assertEqual(self.client.get(self._url("pdf")).status_code, 400)
        self.assertEqual(
            self.client.get(self._url("xlsx", periodo="custom", desde="2026-09-05", hasta="2026-01-01")).status_code,
            400,
        )

    def test_sin_capacidad_de_exportar_devuelve_403(self):
        grupo = Group.objects.create(name="Becas — Reportes solo ver")
        RolMeta.objects.create(grupo=grupo, programa=programa_becas(), activo=True)
        content_type = ContentType.objects.get_for_model(Capacidad)
        for codigo in ("becas.reportes.ver", "becas.segmento.ver"):
            grupo.permissions.add(Permission.objects.get(content_type=content_type, codename=rbac.codename_de(codigo)))
        usuario = User.objects.create_user("solo-ver-dash", password="x")
        usuario.groups.add(grupo)
        AsignacionCoordinador.objects.create(segmento=self.segmento, coordinador=usuario)
        self.client.force_login(usuario)

        self.assertEqual(
            self.client.get(reverse("becas:programa_dashboard_datos", args=[self.programa.pk])).status_code, 200
        )
        self.assertEqual(self.client.get(self._url("xlsx")).status_code, 403)
        pantalla = self.client.get(reverse("becas:programa_detalle", args=[self.programa.pk]))
        self.assertContains(pantalla, 'id="tab-dashboard"')
        self.assertNotContains(pantalla, reverse("becas:programa_dashboard_exportar", args=[self.programa.pk, "xlsx"]))


class DatosConFormasRarasTests(DashboardBecasBase):
    """Filas viejas o de otros canales pueden traer el JSON con otra forma: no tiran el tablero."""

    def test_respuesta_de_tolera_string_dicts_y_bolsas_invalidas(self):
        self.assertEqual(svc.respuesta_de('{"globales": {"7": "Trabaja"}}', "global:7"), ["Trabaja"])
        self.assertEqual(
            svc.respuesta_de({"globales": {"7": {"valor": "Estudia", "etiqueta": "Estudia"}}}, "global:7"), ["Estudia"]
        )
        self.assertEqual(svc.respuesta_de({"globales": ["no", "es", "dict"]}, "global:7"), [])
        self.assertEqual(svc.respuesta_de({"globales": {"7": [{"valor": "A"}, "B", {}]}}, "global:7"), ["A", "B"])
        self.assertEqual(svc.respuesta_de("basura no json", "global:7"), [])
        self.assertEqual(svc.respuesta_de(12, "global:7"), [])

    def test_opciones_con_formas_viejas_se_normalizan(self):
        self.assertEqual(svc._opciones_texto(["A", "B", "A", ""]), ["A", "B"])
        self.assertEqual(svc._opciones_texto('["Sí", "No"]'), ["Sí", "No"])
        self.assertEqual(svc._opciones_texto("Uno\nDos\n"), ["Uno", "Dos"])
        self.assertEqual(svc._opciones_texto([{"valor": "x", "etiqueta": "Equis"}, {"value": "y"}]), ["Equis", "y"])
        self.assertEqual(svc._opciones_texto({"a": "Alfa", "b": "Beta"}), ["Alfa", "Beta"])
        self.assertEqual(svc._opciones_texto(None), [])
        self.assertEqual(svc._opciones_texto(42), [])

    def test_formularios_con_data_rara_no_rompen_las_respuestas(self):
        dentro = HOY - timedelta(days=2)
        q1 = str(self.q_laboral.pk)
        self._formulario(self.rel_propio, creado=dentro, data={"globales": {q1: "Trabaja"}, "requisitos": {}})
        self._formulario(self.rel_propio, creado=dentro, data='{"globales": {"%s": "Estudia"}}' % q1)
        self._formulario(self.rel_propio, creado=dentro, data={"globales": [1, 2, 3], "requisitos": None})
        self._formulario(self.rel_propio, creado=dentro, data=[])  # lista en vez de dict: forma inválida pero no nula
        PreguntaGlobal.objects.filter(pk=self.q_transporte.pk).update(
            opciones=[{"valor": "bus", "etiqueta": "Colectivo"}, "A pie"]
        )

        laboral = svc.distribucion_respuestas(self.admin, self.programa, VENTANA, f"global:{q1}")
        self.assertEqual(laboral.base, 2)
        transporte = next(
            p
            for p in svc.preguntas_graficables(self.admin, self.programa)
            if p.clave == f"global:{self.q_transporte.pk}"
        )
        self.assertEqual(transporte.opciones, ["Colectivo", "A pie"])
        datos = svc.metricas(self.admin, self.programa, VENTANA)
        self.assertEqual(datos.indicadores.formularios_recibidos, 4)

    def test_endpoint_degrada_si_falla_solo_la_pregunta(self):
        from unittest import mock

        self._formulario(self.rel_propio, creado=HOY - timedelta(days=1))
        self.client.force_login(self.admin)
        with mock.patch(
            "programas.services.dashboard_becas.distribuciones_respuestas", side_effect=TypeError("unhashable")
        ):
            respuesta = self.client.get(
                reverse("becas:programa_dashboard_datos", args=[self.programa.pk]) + "?periodo=90"
            )
        self.assertEqual(respuesta.status_code, 200)
        cuerpo = respuesta.json()
        self.assertEqual(cuerpo["datos"]["indicadores"]["formularios_recibidos"], 1)
        self.assertIsNone(cuerpo["respuestas"])
        self.assertEqual(len(cuerpo["avisos"]), 1)
        self.assertIn("TypeError", cuerpo["avisos"][0])

    def test_endpoint_informa_la_etapa_si_fallan_las_metricas(self):
        from unittest import mock

        self.client.force_login(self.admin)
        with mock.patch("programas.services.dashboard_becas.metricas", side_effect=ValueError("boom")):
            respuesta = self.client.get(
                reverse("becas:programa_dashboard_datos", args=[self.programa.pk]) + "?periodo=90"
            )
        self.assertEqual(respuesta.status_code, 500)
        self.assertIn("las métricas", respuesta.json()["errores"][0])
        self.assertIn("ValueError", respuesta.json()["errores"][0])
        self.assertNotIn("boom", respuesta.json()["errores"][0])
